# Standard Library
import csv
import logging
import re
import tempfile

from flask_weasyprint import render_pdf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ...common import *


@blueprint.route("/race/<race_id>/results/export/csv")
@login_required
def race_results_export_csv(race_id):
    race = session.get(db.Race, race_id)
    path = os.path.join(tempfile.mkdtemp(), race.slug + "-results.csv")

    results = db.filter_race_results(session.query(db.Entry), race_id)

    categories = db.entries_get_categories(results)

    # Sort results in each category.
    for results in categories.values():
        results.sort(key=lambda x: x.time, reverse=False)

    with open(path, "w", newline="") as fid:
        writer = csv.writer(fid, quoting=csv.QUOTE_NONNUMERIC)
        for results in categories.values():
            for position, result in enumerate(results):
                writer.writerow(
                    [result.category, position + 1, str(result), result.time]
                )

    return send_file(path, as_attachment=True)


@blueprint.route("/race/<race_id>/results/export/pdf", methods=("GET", "POST"))
@login_required
def race_results_export_pdf(race_id):
    race = session.get(db.Race, race_id)

    if request.method == "POST":
        type = request.form["type"].capitalize()

        return render_pdf(
            url_for("paddle.race_results_paginated", race_id=race.id, type=type),
            download_filename=race.slug + "-results.pdf",
        )

    return render_template(
        "race-results-export-pdf.j2",
        race=race,
    )


@blueprint.route("/race/<race_id>/results/export/xlsx")
@login_required
def race_results_export_xlsx(race_id):
    logging.info("Export XLSX results.")
    race = session.get(db.Race, race_id)
    path = os.path.join(tempfile.mkdtemp(), race.slug + "-results.xlsx")

    results = db.filter_race_results(session.query(db.Entry), race_id)

    categories = db.entries_get_categories(results)

    # Sort results in each category.
    for results in categories.values():
        results.sort(key=lambda x: x.time, reverse=False)

    workbook = Workbook()
    font_head = Font(name="Arial", size=10, color="FFFFFF")
    font_data = Font(name="Arial", size=10, color="000000")

    # Delete default sheet.
    workbook.remove_sheet(workbook.get_sheet_by_name("Sheet"))

    for category, results in categories.items():
        # Create sheet for category.
        category = re.sub(" +", "-", category)
        category = re.sub("/", "+", category)
        sheet = workbook.create_sheet(category)

        # Set column widths.
        sheet.column_dimensions["A"].width = 6  # Number
        sheet.column_dimensions["B"].width = 20  # Surname
        sheet.column_dimensions["C"].width = 20  # First name
        sheet.column_dimensions["D"].width = 9  # BCU number
        sheet.column_dimensions["E"].width = 10  # Expiry
        sheet.column_dimensions["F"].width = 5  # Club
        sheet.column_dimensions["G"].width = 5  # Class
        sheet.column_dimensions["H"].width = 5  # Div
        sheet.column_dimensions["I"].width = 5  # Due
        sheet.column_dimensions["J"].width = 5  # Paid
        sheet.column_dimensions["K"].width = 7  # Start
        sheet.column_dimensions["L"].width = 7  # Finish
        sheet.column_dimensions["M"].width = 4  # Adjustment
        sheet.column_dimensions["N"].width = 7  # Elapsed
        sheet.column_dimensions["O"].width = 7  # Position
        sheet.column_dimensions["P"].width = 5  # Points
        sheet.column_dimensions["Q"].width = 5  # P/D

        # Add header records.
        head = [
            "Number",
            "Surname",
            "First name",
            "BC Number",
            "Expiry",
            "Club",
            "Class",
            "Div",
            "Due",
            "Paid",
            "Start",
            "Finish",
            "Adj",
            "Elapsed",
            "Position",
            "Points",
            "P/D",
        ]
        sheet.append(head)

        # Add results.
        for position, result in enumerate(results):
            for crew in result.crews:
                row = [
                    int(result.race_number),
                    crew.paddler.last,
                    crew.paddler.first,
                    crew.paddler.membership_number,
                    (
                        crew.paddler.membership_expiry.strftime("%d/%m/%Y")
                        if crew.paddler.membership_expiry
                        else None
                    ),
                    crew.club.code if crew.club else None,
                    None,
                    crew.paddler.division,
                    None,
                    None,
                    result.time_start,
                    result.time_finish,
                    result.time_adjustment,
                    result.time,
                    position + 1,
                    None,
                    None,
                ]
                sheet.append(row)

        for cells in sheet.iter_rows(max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in cells:
                if cell.row == 1:
                    cell.font = font_head
                    # Style for header row.
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = PatternFill(
                        start_color="333333", end_color="333333", fill_type="solid"
                    )
                else:
                    cell.font = font_data
                    # Styles for rest of rows.
                    if cell.column_letter in ("E", "K", "L", "N"):
                        cell.alignment = Alignment(horizontal="right")
                    if cell.column_letter in ("F", "H", "O"):
                        cell.alignment = Alignment(horizontal="center")

    workbook.save(path)

    return send_file(path, as_attachment=True)
